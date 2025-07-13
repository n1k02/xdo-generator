import os
import re
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
    get_column_letter
)
import win32com.client as win32 


models_folder_name = "."

def create_xdo_metadata_sheet(wb):
    sheet = wb.create_sheet("XDO_METADATA")

    headers = [
        ("Version", ""),
        ("Extractor Version", ""),
        ("Template Code", ""),
        ("Template Type", "TYPE_EXCEL_TEMPLATE"),
        ("Preprocess XSLT File", ""),
        ("Last Modified Date", ""),
        ("Last Modified By", ""),
        ("", ""),
        ("Data Constraints:", "")
    ]

    for i, (key, value) in enumerate(headers):
        sheet.cell(row=i + 1, column=1).value = key
        sheet.cell(row=i + 1, column=2).value = value

    return sheet


def find_tagged_cells(file_path):
    wb = load_workbook(file_path)
    tag_pattern = re.compile(r'^G(\d)(\d{2})$')  # G1XX, G2XX, etc.
    model_fields = defaultdict(list)

    for sheet in wb.worksheets:
        if sheet.title == "XDO_METADATA":
            continue

        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if isinstance(cell_value, str):
                    match = tag_pattern.match(cell_value.strip())
                    if match:
                        model_number = match.group(1)
                        field_number = match.group(2)
                        model_key = f'G{model_number}'
                        global_field_num = int(model_number + field_number)
                        model_fields[model_key].append(global_field_num)

    return model_fields


def split_sql_fields(sql):
    fields = []
    current = ''
    depth = 0

    for char in sql:
        if char == ',' and depth == 0:
            fields.append(current.strip())
            current = ''
        else:
            current += char
            if char == '(':
                depth += 1
            elif char == ')':
                depth -= 1
    if current:
        fields.append(current.strip())

    return fields

def extract_field_alias(field_expr):
    as_match = re.search(r'\s+AS\s+("?\w+"?)$', field_expr, re.IGNORECASE)
    if as_match:
        return as_match.group(1).replace('"', '')

    dot_match = re.search(r'(\w+)\.(\w+)$', field_expr)
    if dot_match:
        return dot_match.group(2)

    tokens = re.findall(r'\w+', field_expr)
    if tokens:
        return tokens[-1]

    return 'UNKNOWN'

def parse_field_names_from_txt(file_path):
    with open(file_path, encoding='utf-8') as f:
        content = f.read()

    field_expressions = split_sql_fields(content)
    parsed_fields = [extract_field_alias(expr) for expr in field_expressions]
    return parsed_fields


def fill_metadata_body(sheet, model_fields_dict, base_row=10):
    current_row = base_row

    for group_name in sorted(model_fields_dict.keys()):
        model_index = int(group_name[1:])
        tag_ids = sorted(model_fields_dict[group_name])

        txt_file_path = f"{models_folder_name}/{group_name}.txt"
        field_names = parse_field_names_from_txt(txt_file_path)

        if len(field_names) < len(tag_ids):
            print(f"⚠️ Warning: Not enough fields in {txt_file_path} for {group_name}")
        elif len(field_names) > len(tag_ids):
            print(f"⚠️ Warning: More fields in {txt_file_path} than tags in Excel for {group_name}")

        for i, tag_id in enumerate(tag_ids):
            if i >= len(field_names):
                break
            field = field_names[i]
            tag = f"XDO_?XDOFIELD{tag_id}?"
            value = f"<?{field.upper()}?>"
            sheet.cell(row=current_row, column=1).value = tag
            sheet.cell(row=current_row, column=2).value = value
            current_row += 1

        group_tag = f"XDO_GROUP_?XDOG{model_index}?"
        loop_value = f"<xsl:for-each select=\".//G_{model_index}\">"
        sheet.cell(row=current_row, column=1).value = group_tag
        sheet.cell(row=current_row, column=2).value = loop_value
        sheet.cell(row=current_row, column=3).value = "</xsl:for-each>"
        current_row += 1



def assign_named_ranges(wb):
    tag_pattern = re.compile(r'^G(\d)(\d{2})$')
    grouped_cells = defaultdict(list)

    for sheet in wb.worksheets:
        if sheet.title == "XDO_METADATA":
            continue

        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    match = tag_pattern.fullmatch(cell.value.strip())
                    if match:
                        model = match.group(1)
                        field_num = match.group(2)
                        tag_name = f"XDO_?XDOFIELD{model}{field_num}?"
                        group_key = f"G{model}"

                        cell_ref = f"{sheet.title}!${cell.column_letter}${cell.row}"

                        defined_name = DefinedName(name=tag_name, attr_text=cell_ref)
                        wb.defined_names.add(defined_name)

                        grouped_cells[group_key].append(cell_ref)

    for group_name, refs in grouped_cells.items():
        model_index = group_name[1:]
        group_tag = f"XDO_GROUP_?XDOG{model_index}?"

        rows = []
        cols = []

        for ref in refs:
            sheet_name, coord = ref.split("!")
            coord = coord.replace("$", "")
            col_letter, row = coordinate_from_string(coord)
            col_idx = column_index_from_string(col_letter)
            rows.append(int(row))
            cols.append(col_idx)

        min_col = get_column_letter(min(cols))
        max_col = get_column_letter(max(cols))
        min_row = min(rows)
        max_row = max(rows)

        range_ref = f"{sheet_name}!${min_col}${min_row}:${max_col}${max_row}"

        defined_name = DefinedName(name=group_tag, attr_text=range_ref)
        wb.defined_names.add(defined_name)


def convert_xlsx_to_xls(input_path, output_path):
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except PermissionError:
            print(f"❌ Cannot delete file {output_path}. It may be open.")
            return

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(os.path.abspath(input_path))
    wb.SaveAs(os.path.abspath(output_path), FileFormat=56)
    wb.Close(False)
    excel.Quit()
    print(f"✅ Saved as .xls: {output_path}")


def find_tagged_sheets(wb):
    tag_pattern = re.compile(r'^G(\d)(\d{2})$')
    tagged_sheets = set()

    for sheet in wb.worksheets:
        if sheet.title == "XDO_METADATA":
            continue
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if isinstance(cell_value, str) and tag_pattern.match(cell_value.strip()):
                    tagged_sheets.add(sheet.title)
                    break
    return tagged_sheets

def find_tagged_cells_from_workbook(wb):
    tag_pattern = re.compile(r'^G(\d)(\d{2})$')
    model_fields = defaultdict(list)

    for sheet in wb.worksheets:
        if sheet.title == "XDO_METADATA":
            continue

        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if isinstance(cell_value, str):
                    match = tag_pattern.match(cell_value.strip())
                    if match:
                        model_number = match.group(1)
                        field_number = match.group(2)
                        model_key = f'G{model_number}'
                        global_field_num = int(model_number + field_number)
                        model_fields[model_key].append(global_field_num)

    return model_fields

def find_first_empty_row_after_constraints(sheet):
    row_with_constraints = None

    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=1).value
        if isinstance(value, str) and "Data Constraints:" in value:
            row_with_constraints = row
            break

    if row_with_constraints is None:
        raise ValueError("⛔ 'Data Constraints:' row not found in XDO_METADATA sheet")

    row = row_with_constraints + 1
    while sheet.cell(row=row, column=1).value is not None:
        row += 1

    return row


def main():
    template_path = "template.xlsx"
    wb = load_workbook(template_path)

    tagged_sheets = find_tagged_sheets(wb)

    for sheet in wb.worksheets:
        if sheet.title in tagged_sheets and ' ' in sheet.title:
            old_title = sheet.title
            new_title = sheet.title.replace(' ', '_')
            sheet.title = new_title
            print(f"ℹ️ Renamed sheet: '{old_title}' → '{new_title}'")

    model_fields_dict = find_tagged_cells_from_workbook(wb)

    print("🔍 Found models and fields:")
    for group, tags in model_fields_dict.items():
        print(f"  {group}: {tags}")

    if "XDO_METADATA" in wb.sheetnames:
        sheet = wb["XDO_METADATA"]
        next_row = find_first_empty_row_after_constraints(sheet)
        if next_row is not None:
            metadata_sheet = sheet
            print("ℹ️ Found 'Data Constraints:' in XDO_METADATA. Appending new entries.")
        else:
            del wb["XDO_METADATA"]
            metadata_sheet = create_xdo_metadata_sheet(wb)
            print("⚠️ 'Data Constraints:' not found in XDO_METADATA. Sheet recreated.")
            next_row = 11
    else:
        metadata_sheet = create_xdo_metadata_sheet(wb)
        print("ℹ️ XDO_METADATA sheet created from scratch.")
        next_row = 11

    fill_metadata_body(metadata_sheet, model_fields_dict, base_row=next_row)
    assign_named_ranges(wb)

    xlsx_path = "temp.xlsx"
    xls_path = "template.xls"

    try:
        if os.path.exists(xlsx_path):
            os.remove(xlsx_path)
        wb.save(xlsx_path)
    except PermissionError:
        print(f"❌ Cannot delete or save file {xlsx_path}. It may be open.")
        return

    convert_xlsx_to_xls(xlsx_path, xls_path)

    try:
        os.remove(xlsx_path)
    except PermissionError:
        print(f"❌ Could not delete temporary file {xlsx_path}. It may be open.")

    print(f"✅ Done: {xls_path}")


if __name__ == "__main__":
    main()
